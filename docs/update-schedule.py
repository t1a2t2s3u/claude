# SEO/MEO 更新スケジュールのスプレッドシートを生成する。
# 配色はサイト (site/assets/style.css) と同じ。運用者は開発者ではないため、
# シート名・列名・本文はすべて日本語にする。

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BRAND_DEEP = "0B4A5F"
BRAND = "00A3C9"
ACCENT = "E2661F"
LIGHT = "EAF4F7"      # 薄い水色（交互行）
NOTE = "FFF6E8"       # 薄いオレンジ（注記・記入欄）
FONT = "Meiryo"

wb = Workbook()

thin = Side(style="thin", color="BBCDD4")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=BRAND_DEEP)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

def style_body(ws, r1, r2, ncols, zebra=True):
    for r in range(r1, r2 + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            if cell.font.name != FONT or not cell.font.bold:
                if cell.font.color is None or cell.font.color.rgb in (None, "FF000000"):
                    cell.font = Font(name=FONT, size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if zebra and r % 2 == 0:
                if cell.fill.fgColor.rgb in (None, "00000000"):
                    cell.fill = PatternFill("solid", fgColor=LIGHT)

# ---------------------------------------------------------------- シート1
ws = wb.active
ws.title = "更新スケジュール"

ws["A1"] = "辰弥塗装工業　SEO・MEO 更新スケジュール"
ws["A1"].font = Font(name=FONT, bold=True, size=14, color=BRAND_DEEP)
ws["A2"] = "青字のセル（1回の時間・月の回数）は目安です。実態に合わせて書き換えると、右の「月の合計時間」が自動で計算し直されます。"
ws["A2"].font = Font(name=FONT, size=9, color="555555")
ws["A3"] = "「工事完了ごと」の回数は、月2件の工事を仮定しています（ご本人からの情報ではなく仮置きの数字です）。"
ws["A3"].font = Font(name=FONT, size=9, color="555555")

headers = ["媒体", "やること", "頻度", "タイミングの目安",
           "1回の時間(分)", "月の回数(目安)", "月の合計(分)", "優先度",
           "ねらい（SEO・MEO上の理由）", "やり方のメモ"]
HR = 5
for i, h in enumerate(headers, 1):
    ws.cell(row=HR, column=i, value=h)
style_header(ws, HR, len(headers))

rows = [
    ("Google ビジネスプロフィール", "口コミへの返信", "来たら都度", "48時間以内", 10, 2, "★★★",
     "返信率と返信の早さはMEOの評価要素。読んでいる見込み客への信頼材料にもなる",
     "定型文にせず、その工事の内容に一言触れて返す"),
    ("Google ビジネスプロフィール", "口コミの依頼", "工事完了ごと", "引き渡しのとき", 5, 2, "★★★",
     "口コミの「数」と「新しさ」はMEOで最も効く要素。依頼しないと増えない",
     "引き渡し時に口頭でお願いし、QRコードの紙を渡す"),
    ("Google ビジネスプロフィール", "投稿（最新情報）", "週1回", "曜日を決めて固定（例：金曜の夕方）", 15, 4, "★★★",
     "プロフィールが動いている店として扱われ、検索結果での見え方が良くなる",
     "現場写真1枚＋2〜3行で十分。Instagramと同じネタの使い回しでよい"),
    ("Google ビジネスプロフィール", "写真の追加", "週1回（投稿と同時）", "投稿と同時", 5, 4, "★★☆",
     "写真の多いプロフィールは閲覧・経路検索が増える傾向がある",
     "施工中・道具・人が写っている写真が強い"),
    ("Google ビジネスプロフィール", "登録情報の点検", "3か月に1回", "3・6・9・12月の月初", 15, 0.33, "★★☆",
     "サイト(company.toml)と住所・電話・営業時間がずれると減点要因になる",
     "営業時間・サービス提供地域・ウェブサイト欄を見る。年末年始などの特別営業時間も"),
    ("自社サイト", "施工事例の追加", "工事完了ごと", "完了から1週間以内", 60, 2, "★★★",
     "問い合わせに最も直結するページ。地名入りの事例が「地名＋外壁塗装」検索の受け皿になる",
     "before/after写真を現場で撮っておくのが前提。ページ作成はClaudeに任せて確認だけでもよい"),
    ("自社サイト", "ブログ記事", "月2本", "月の前半・後半に1本ずつ", 90, 2, "★★☆",
     "検索の入り口を増やす。本数より、疑問にきちんと答える質のほうが効く",
     "「ブログネタ帳」シート参照。下書きはClaudeに依頼し、内容の確認に集中する。確認が負担な月は1本に落としてよい"),
    ("自社サイト", "既存記事の見直し", "年2回", "4月（補助金記事）と10月", 30, 0.17, "★★☆",
     "古い数値・年度の残った記事は信頼と順位を落とす。特に補助金記事は毎年度変わる",
     "補助金の額・受付状況、料金、リンク切れを確認"),
    ("Instagram", "フィード投稿", "週1〜2回", "現場のある日", 15, 6, "★☆☆",
     "検索順位への直接効果は薄い。会社名で調べた人の信頼材料＋GBP投稿のネタ元として",
     "現場写真中心。ハッシュタグは #秋田市外壁塗装 など地名入りを固定で"),
    ("Instagram", "ストーリーズ", "現場のある日に随時", "作業の合間", 3, 8, "★☆☆",
     "手間最小で「動いている会社」だと伝わる。無理のない範囲で",
     "作業風景をそのまま。編集しない"),
    ("公式LINE", "一斉配信", "年2回", "1〜2月と7〜8月（工事シーズン前）", 30, 0.17, "★★☆",
     "シーズン前の思い出しに効く。配信しすぎはブロックの元（docs/line.md 参照）",
     "ネタはブログ記事の流用でよい"),
    ("分析", "週次レポートの確認", "週1回", "月曜の朝", 5, 4, "★★☆",
     "GBPの表示回数・電話・経路検索の動きを見て、効いている施策を知る",
     "seo-meo report を実行（Claudeに頼んでもよい）"),
    ("分析", "Search Console の確認", "月1回", "月初", 15, 1, "★☆☆",
     "実際に検索された言葉が分かる。ブログと施工事例のネタに反映する",
     "サイトをSearch Consoleに登録してから（docs/deploy.md 手順5）"),
]

r = HR + 1
first_data = r
for row in rows:
    media, task, freq, timing, minutes, times, prio, why, memo = row
    ws.cell(row=r, column=1, value=media)
    ws.cell(row=r, column=2, value=task).font = Font(name=FONT, bold=True, size=10)
    ws.cell(row=r, column=3, value=freq)
    ws.cell(row=r, column=4, value=timing)
    c5 = ws.cell(row=r, column=5, value=minutes)
    c6 = ws.cell(row=r, column=6, value=times)
    for c in (c5, c6):
        c.font = Font(name=FONT, size=10, color="0000FF")
        c.alignment = Alignment(horizontal="center", vertical="top")
    c6.number_format = "0.##"
    c7 = ws.cell(row=r, column=7, value=f"=E{r}*F{r}")
    c7.number_format = "0"
    c7.alignment = Alignment(horizontal="center", vertical="top")
    c8 = ws.cell(row=r, column=8, value=prio)
    c8.alignment = Alignment(horizontal="center", vertical="top")
    if prio == "★★★":
        c8.font = Font(name=FONT, bold=True, size=10, color=ACCENT)
    ws.cell(row=r, column=9, value=why)
    ws.cell(row=r, column=10, value=memo)
    r += 1
last_data = r - 1

# 合計行
ws.cell(row=r, column=2, value="月の作業時間の合計").font = Font(name=FONT, bold=True, size=10)
tc = ws.cell(row=r, column=7, value=f"=SUM(G{first_data}:G{last_data})")
tc.number_format = "0"
tc.font = Font(name=FONT, bold=True, size=10)
tc.alignment = Alignment(horizontal="center")
hc = ws.cell(row=r, column=9, value=f"=\"1か月あたり約 \"&ROUND(G{r}/60,1)&\" 時間（1日あたり10分程度）\"")
hc.font = Font(name=FONT, bold=True, size=10, color=BRAND_DEEP)
for c in range(1, 11):
    ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=NOTE)
    ws.cell(row=r, column=c).border = border

style_body(ws, first_data, last_data, 10)

widths = [22, 18, 14, 22, 9, 9, 9, 8, 40, 36]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = f"A{HR+1}"
ws.row_dimensions[HR].height = 28

# ---------------------------------------------------------------- シート2
ws2 = wb.create_sheet("年間カレンダー")
ws2["A1"] = "年間カレンダー　— 秋田の塗装シーズン（4〜11月）に合わせた重点"
ws2["A1"].font = Font(name=FONT, bold=True, size=13, color=BRAND_DEEP)

headers2 = ["月", "時期の特徴", "その月の重点", "ブログ・投稿・配信のネタ例"]
HR2 = 3
for i, h in enumerate(headers2, 1):
    ws2.cell(row=HR2, column=i, value=h)
style_header(ws2, HR2, len(headers2))

calendar = [
    ("1月", "雪・閑散期", "公式LINE 一斉配信①（春の工事の案内）。春の予約を受け始める", "「春に塗りたいなら今から動く」— 塗り替え時期の記事を再利用"),
    ("2月", "雪・閑散期", "ブログの仕込み（現場が無い時期にまとめて下書き）", "塗料の選び方、見積書の見方など、季節を問わないネタ"),
    ("3月", "雪解け", "補助金記事を新年度版に更新する準備（市・県の発表を確認）", "「今年の補助金はどうなる？」の下調べ"),
    ("4月", "シーズン開始。市の補助金受付が例年開始", "補助金記事を新年度の数値に更新して公開。GBP投稿でも告知", "補助金記事（毎年4月に更新）。市の受付は先着順なので早い告知が価値になる"),
    ("5月", "繁忙期（希望が集中）", "現場写真を撮りためる。完工ごとに施工事例＋口コミ依頼", "現場のbefore/after"),
    ("6月", "梅雨入り", "通常運用", "「雨の日、塗装工事はどうなる？」"),
    ("7月", "夏", "公式LINE 一斉配信②（秋の工事の案内）", "「秋の工事は夏のうちに相談を」"),
    ("8月", "繁忙期", "通常運用", "現場のbefore/after、職人の紹介"),
    ("9月", "繁忙期（気候が安定）", "通常運用", "現場のbefore/after"),
    ("10月", "シーズン終盤", "既存記事の見直し（年2回の2回目）。「年内に塗れるか」の相談が増える", "「今年中に間に合う？」— 塗り替え時期の記事へ誘導"),
    ("11月", "完工ラッシュ", "施工事例をまとめて追加。口コミ依頼を忘れずに", "1年の施工まとめ"),
    ("12月", "雪・閑散期", "1年のレポートを振り返り、来年の計画。Search Consoleで1年分の検索語を確認", "「今年もありがとうございました」投稿。雪止め・落雪のネタ"),
]
r = HR2 + 1
for month, season, focus, neta in calendar:
    mc = ws2.cell(row=r, column=1, value=month)
    mc.font = Font(name=FONT, bold=True, size=10, color=BRAND_DEEP)
    mc.alignment = Alignment(horizontal="center", vertical="top")
    ws2.cell(row=r, column=2, value=season)
    ws2.cell(row=r, column=3, value=focus)
    ws2.cell(row=r, column=4, value=neta)
    # 塗装シーズンの月をうっすら色分け
    if month in ("4月", "5月", "6月", "7月", "8月", "9月", "10月", "11月"):
        ws2.cell(row=r, column=1).fill = PatternFill("solid", fgColor="D9EEF5")
    r += 1
style_body(ws2, HR2 + 1, r - 1, 4, zebra=False)
for i, w in enumerate([7, 20, 46, 46], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = f"A{HR2+1}"

# ---------------------------------------------------------------- シート3
ws3 = wb.create_sheet("ブログネタ帳")
ws3["A1"] = "ブログネタ帳　— 優先順位と根拠は docs/keyword-research-2026-08.xlsx（2026年8月のキーワード調査）を参照"
ws3["A1"].font = Font(name=FONT, bold=True, size=13, color=BRAND_DEEP)

headers3 = ["ネタ", "狙う検索の言葉", "おすすめの時期", "状態", "メモ"]
HR3 = 3
for i, h in enumerate(headers3, 1):
    ws3.cell(row=HR3, column=i, value=h)
style_header(ws3, HR3, len(headers3))

topics = [
    ("秋田市の外壁塗装に使える補助金・助成金", "秋田市 外壁塗装 補助金", "4月（毎年更新）", "下書き済み・確認待ち",
     "drafts/ に下書きあり。毎年4月に新年度の数値へ更新する"),
    ("塗料の種類と耐用年数の選び方", "外壁塗装 塗料 種類 / シリコン フッ素 違い", "いつでも", "未着手",
     "services.toml の料金・保証年数と食い違わないように"),
    ("外壁塗装の見積書の見方・比べ方", "外壁塗装 見積り 見方", "いつでも", "未着手",
     "「一式」表記への注意など。相見積り歓迎の姿勢が信頼になる"),
    ("訪問販売で「今すぐ塗らないと危ない」と言われたら", "外壁塗装 訪問販売 しつこい", "いつでも", "未着手",
     "docs/line.md でも受注につながりやすいと分析済みのテーマ"),
    ("外壁塗装の工事の流れ（足場から完工まで）", "外壁塗装 工事 流れ 日数", "いつでも", "未着手",
     "LINEの工期応答の内容を膨らませる。10日前後の工程表"),
    ("コーキング（目地）の役割と劣化のサイン", "コーキング 劣化 ひび割れ", "いつでも", "未着手",
     "施工写真 shokunin-coking.jpg が使える"),
    ("雪止め・屋根の雪対策と塗装", "秋田 屋根 雪止め 落雪", "10〜12月", "未着手",
     "市の補助金でも雪止め設置は対象工事。冬前に出すと効く"),
    ("外壁の色選びのコツと失敗例", "外壁塗装 色見本(2,400) / 人気色(1,900) / 色選び シュミレーション(1,900)", "いつでも", "優先度UP",
     "キーワード調査で需要が想定以上と判明（合計1万超/月）。施工事例の写真がそのまま使える。9月前半の本に"),
    ("よくあるご質問まとめ", "外壁塗装 よくある質問", "いつでも", "未着手",
     "LINEの応答メッセージ（無料・工期・保証・エリア）を記事に再構成"),
    ("火災保険と外壁・屋根の修理の考え方", "屋根 修理 火災保険", "いつでも", "要注意",
     "適用可否は個別判断。「保険で無料」と断定しない書き方が必須（docs/line.md 参照）"),
    ("外壁塗装を20年していないとどうなる？", "外壁塗装 20年してない / 30年してない（知恵袋の質問需要1位）", "9月後半", "優先度UP",
     "悩み系の頂点。凍害・雪という秋田固有の劣化を絡めれば全国サイトに勝てる。「点検だけでもOK」と相性が良い"),
    ("塗装と張り替え（カバー工法）どっちが安い？", "外壁塗装 張り替え どっち / サイディング 張替え", "11月", "未着手",
     "検討初期の比較記事。張替えが向くケースも正直に書くと信頼になる"),
]
r = HR3 + 1
for t in topics:
    for c, v in enumerate(t, 1):
        ws3.cell(row=r, column=c, value=v)
    st = ws3.cell(row=r, column=4)
    st.alignment = Alignment(horizontal="center", vertical="top")
    if t[3] in ("下書き済み・確認待ち", "優先度UP"):
        st.font = Font(name=FONT, bold=True, size=10, color=ACCENT)
    elif t[3] == "要注意":
        st.font = Font(name=FONT, bold=True, size=10, color="B00000")
    r += 1
style_body(ws3, HR3 + 1, r - 1, 5)
for i, w in enumerate([38, 30, 15, 16, 44], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.freeze_panes = f"A{HR3+1}"

# ---------------------------------------------------------------- シート4
ws4 = wb.create_sheet("実績記録")
ws4["A1"] = "実績記録　— やったことを1行ずつ足していくだけ"
ws4["A1"].font = Font(name=FONT, bold=True, size=13, color=BRAND_DEEP)
ws4["A2"] = "続いているかを後から振り返るための記録です。1行目はグレーの記入例（消して使ってください）。"
ws4["A2"].font = Font(name=FONT, size=9, color="555555")

headers4 = ["日付", "媒体", "内容", "リンク・メモ"]
HR4 = 4
for i, h in enumerate(headers4, 1):
    ws4.cell(row=HR4, column=i, value=h)
style_header(ws4, HR4, len(headers4))

example = ["2026/8/26", "GBP投稿", "土崎の外壁工事の完了報告（写真1枚）", "口コミ依頼も済み"]
for c, v in enumerate(example, 1):
    cell = ws4.cell(row=HR4 + 1, column=c, value=v)
    cell.font = Font(name=FONT, size=10, italic=True, color="888888")
    cell.border = border
    cell.alignment = Alignment(vertical="top", wrap_text=True)
# 空の記入行を用意しておく
for r in range(HR4 + 2, HR4 + 32):
    for c in range(1, 5):
        cell = ws4.cell(row=r, column=c)
        cell.border = border
        cell.fill = PatternFill("solid", fgColor="FFFDF7")
for i, w in enumerate([12, 14, 44, 32], 1):
    ws4.column_dimensions[get_column_letter(i)].width = w
ws4.freeze_panes = f"A{HR4+1}"

from pathlib import Path
out = str(Path(__file__).with_name("update-schedule.xlsx"))
wb.save(out)
print("saved:", out)
